"""
Sprint 3 - Day 17
Export Screener Results to Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.screener.engine import run_screener


GREEN_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

HEADER_FILL = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)


PRESETS = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
    "turnaround_watch"
]


DISPLAY_COLUMNS = [
    "company_id",
    "year",
    "composite_quality_score",
    "sector_quality_score",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "asset_turnover",
    "sales",
    "net_profit"
]


def format_sheet(ws):

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    ws.freeze_panes = "A2"

    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except:
                pass

        ws.column_dimensions[
            column_letter
        ].width = max_length + 2


def export_screeners():

    wb = Workbook()

    wb.remove(wb.active)

    for preset in PRESETS:

        df = run_screener(preset)

        ws = wb.create_sheet(
            title=preset[:31]
        )

        ws.append(DISPLAY_COLUMNS)

        for _, row in df.iterrows():

            ws.append(
                [
                    row.get(col)
                    for col in DISPLAY_COLUMNS
                ]
            )

        format_sheet(ws)

    wb.save(
        "output/screener_output.xlsx"
    )

    print(
        "screener_output.xlsx created"
    )


if __name__ == "__main__":

    export_screeners()