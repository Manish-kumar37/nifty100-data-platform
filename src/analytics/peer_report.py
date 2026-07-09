"""
Sprint 3 - Day 20
Peer Comparison Report
"""

import sqlite3
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics.peer import latest_records
from src.analytics.scoring import add_composite_scores

DB_PATH = "nifty100.db"

HEADER_FILL = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

BENCHMARK_FILL = PatternFill(
    start_color="FFD966",
    end_color="FFD966",
    fill_type="solid"
)

DISPLAY_COLUMNS = [
    "company_id",
    "year",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "composite_quality_score",
    "sector_quality_score"
]


def load_data():

    conn = sqlite3.connect(DB_PATH)

    ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    peers = pd.read_sql(
        "SELECT * FROM peer_groups",
        conn
    )

    conn.close()

    return ratios, peers


def auto_width(ws):

    for column in ws.columns:

        letter = get_column_letter(column[0].column)

        width = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column
        )

        ws.column_dimensions[letter].width = width + 2


def generate_report():

    ratios, peers = load_data()

    ratios = add_composite_scores(ratios)

    ratios = latest_records(ratios)

    report = ratios.merge(

        peers[
            [
                "company_id",
                "peer_group_name",
                "is_benchmark"
            ]
        ],

        on="company_id",

        how="inner"

    )

    wb = Workbook()

    wb.remove(wb.active)

    for group in sorted(
        report["peer_group_name"].unique()
    ):

        sheet = wb.create_sheet(group[:31])

        sheet.append(DISPLAY_COLUMNS)

        for cell in sheet[1]:

            cell.fill = HEADER_FILL

            cell.font = HEADER_FONT

        row_number = 2

        group_df = report[
            report["peer_group_name"] == group
        ]

        for _, row in group_df.iterrows():

            values = [
                row.get(col)
                for col in DISPLAY_COLUMNS
            ]

            sheet.append(values)

            if row["is_benchmark"] == 1:

                for cell in sheet[row_number]:

                    cell.fill = BENCHMARK_FILL

            row_number += 1

        # Median Row

        numeric = group_df[
            DISPLAY_COLUMNS[2:]
        ].median(
            numeric_only=True
        )

        median_row = [
            "Median",
            ""
        ] + numeric.tolist()

        sheet.append(median_row)

        auto_width(sheet)

        sheet.freeze_panes = "A2"

    wb.save(
        "output/peer_comparison.xlsx"
    )

    print()

    print("peer_comparison.xlsx created")

    print(
        "Sheets:",
        len(wb.sheetnames)
    )


if __name__ == "__main__":

    generate_report()